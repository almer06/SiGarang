from datetime import date, timedelta, datetime

import pandas as pd
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.contrib.auth.views import LoginView, LogoutView, PasswordResetView, PasswordResetConfirmView
from django.contrib.sites.shortcuts import get_current_site
from django.core.paginator import Paginator
from django.db import connection
from django.http import HttpResponseForbidden, HttpResponseRedirect, HttpRequest, JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.generic import ListView, DetailView, CreateView, TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.child import INVALID_TITLE_REGEX

from administrator.models import (UKM, IKM, Agenda, Document, UnitGroceries,
                                  VariantGroceries, Market, KiosPupuk, AgenLPG, StockItem)
from visitor.forms import CustomUserCreationForm, SendInBluePasswordResetForm
from visitor.models import User
from visitor.sendinblue import SendInBlue
from visitor.tokens import activation_account


class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    login_url = reverse_lazy('visitor:login')
    permission_denied_message = 'Access Denied.'
    request = HttpRequest  # Testing

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        return HttpResponseForbidden("You do not have permission to access this page.")


class AgendaVisitor(ListView):
    template_name = 'visitor/agenda.html'
    model = Agenda
    paginate_by = 10
    ordering = ['agenda_created']
    queryset = Agenda.objects.filter(agenda_is_publish=True)
    extra_context = {
        'title': 'Agenda'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(agenda_name__icontains=name_filter)
        return queryset


class UKMVisitor(ListView):
    template_name = 'visitor/ukm.html'
    model = UKM
    paginate_by = 12
    extra_context = {
        'title': 'UKM'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(ukm_name__icontains=name_filter)
        return queryset


class UKMDetailVisitor(DetailView):
    model = UKM
    template_name = 'visitor/ukm-detail.html'
    slug_field = 'ukm_name_slug'
    extra_context: dict = {}

    def get(self, request, *args, **kwargs):
        self.extra_context.update({'title': kwargs.get('slug')})
        return super().get(request)


class IKMVisitor(ListView):
    template_name = 'visitor/ikm.html'
    model = IKM
    paginate_by = 12
    extra_context = {
        'title': 'IKM'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(ikm_name__icontains=name_filter)
        return queryset


class IKMDetailVisitor(DetailView):
    model = IKM
    template_name = 'visitor/ikm-detail.html'
    slug_field = 'ikm_name_slug'
    extra_context: dict = {}

    def get(self, request, *args, **kwargs):
        self.extra_context.update({'title': kwargs.get('slug')})
        return super().get(request)


class DocumentVisitor(ListView):
    model = Document
    template_name = 'visitor/document.html'
    extra_context = {
        'title': 'Dokumen'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name_document')
        if name_filter:
            queryset = queryset.filter(document_name__icontains=name_filter)
        return queryset


class HargaSembako(TemplateView):
    template_name = 'visitor/harga-sembako.html'
    extra_context = {
        'title': 'Harga Sembako'
    }


class LoginVisitor(LoginView):
    template_name = 'visitor/login.html'
    redirect_authenticated_user = True
    extra_context = {
        'title': 'Login'
    }
    success_url = reverse_lazy('visitor:harga_sembako')
    next_page = reverse_lazy('visitor:harga_sembako')


class RegisterVisitor(UserPassesTestMixin, CreateView):
    form_class = CustomUserCreationForm
    model = User
    template_name = 'visitor/registration.html'
    success_url = reverse_lazy('visitor:login')
    label_suffix = ' :'
    login_url = reverse_lazy('visitor:document')
    user = None
    object = None
    send_in_blue = SendInBlue

    def get_label_suffix(self):
        return self.label_suffix

    def get_form_kwargs(self):
        kwargs = {
            "initial": self.get_initial(),
            "prefix": self.get_prefix(),
            "label_suffix": self.get_label_suffix()
        }
        if self.request.method in ("POST", "PUT"):
            kwargs.update(
                {
                    "data": self.request.POST,
                    "files": self.request.FILES,
                }
            )
        return kwargs

    def test_func(self):
        return not self.request.user.is_authenticated

    def handle_no_permission(self):
        return HttpResponseRedirect(self.login_url)

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        self.object = None
        if form.is_valid():
            self.user = form.save()
            self.prepare_send_token()
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def prepare_send_token(self):
        current_site = get_current_site(self.request)
        site_name = current_site.name
        domain = current_site.domain
        use_https = self.request.is_secure()

        context = {
            'uid': urlsafe_base64_encode(force_bytes(self.user.pk)),
            'site_name': site_name,
            'domain': domain,
            'token': activation_account.make_token(self.user),
            'protocol': 'https' if use_https else 'http'
        }

        self.get_send_in_blue().email_sendinblue(
            'visitor/activation_user_subject.txt',
            'visitor/activation_user_email.html',
            context,
            from_email=None,
            to_email=self.user.user_email,
        )

    def get_send_in_blue(self):
        sendinblue = self.send_in_blue
        return sendinblue()


class LogoutVisitor(LogoutView):
    next_page = reverse_lazy('visitor:harga_sembako')


class PasswordResetViewVisitor(PasswordResetView):
    success_url = reverse_lazy('visitor:reset_password')
    form_class = SendInBluePasswordResetForm
    email_template_name = 'visitor/email_forget.html'
    template_name = 'visitor/password_reset_form.html'

    def form_valid(self, form):
        messages.success(self.request, "Email reset kata sandi sudah dikirim. Cek inbox/spam Anda. Terima kasih!")
        return super().form_valid(form)


class PasswordResetConfirmViewVisitor(PasswordResetConfirmView):
    success_url = reverse_lazy('visitor:password_reset_complete')
    template_name = 'visitor/password_reset_confirm.html'


class ActivateAccount(View):
    http_method_names = ['get']

    def get(self, request, *args, **kwargs):
        uid = force_str(urlsafe_base64_decode(kwargs['uidb64']))
        try:
            user = User.objects.get(user_id=uid)
        except User.DoesNotExist:
            return HttpResponse('Not found')

        valid_token = activation_account.check_token(user, kwargs['token'])

        if not valid_token:
            messages.error(request, 'Link sudah tidak bisa digunakan')
            return redirect(reverse_lazy('visitor:login'))

        user.is_active = True
        user.save()

        messages.success(request, 'Account anda telah diaktivasi')
        return redirect(reverse_lazy('visitor:login'))


def sembako(request):
    page = request.GET.get('page')
    page_size = request.GET.get('page_size', 10)

    id_sembako = request.GET.get('id_sembako')
    create_day = request.GET.get('create')
    name = request.GET.get('nama_sembako')

    if id_sembako is None and create_day is None:
        if name:
            query = (VariantGroceries.objects.filter(groceries_name__icontains=name)
                     .values('groceries_id', 'groceries_name', 'groceries_massa', 'groceries_quantity')
                     .order_by('-groceries_created'))
        else:
            query = VariantGroceries.objects.all().values('groceries_id', 'groceries_name', 'groceries_massa',
                                                          'groceries_quantity').order_by('-groceries_created')

        # Create a Paginator
        paginator = Paginator(query, page_size)

        try:
            paginated_data = paginator.page(page)
        except Exception as e:
            paginated_data = paginator.page(1)  # Default to the first page if page is invalid

        return JsonResponse(
            {
                'success': True,
                'data': list(paginated_data),
                'page': paginated_data.number,
                'total_pages': paginator.num_pages
            }
        )

    if id_sembako is None:
        return JsonResponse(
            {
                'success': False,
                'message': 'id is required'
            },
            status=400
        )

    if create_day is None:
        create_day = 'today'

    if create_day not in ['today', 'yesterday']:
        return JsonResponse(
            {
                'success': False,
                'message': 'value must be today or yesterday'
            },
            status=400
        )

    day = None
    if create_day == 'today':
        day = date.today()

    if create_day == 'yesterday':
        day = date.today() - timedelta(days=1)

    query = UnitGroceries.objects.filter(unit_groceries_variant__groceries_id=id_sembako) \
        .filter(unit_groceries_created=day) \
        .values('unit_groceries_price', )

    return JsonResponse(
        {
            'success': True,
            'data': list(query)
        },
        safe=False
    )


def export_excel_harga_sembako(request):
    tanggal = datetime.now().strftime("%d/%B/%Y")
    name_excel = f"harga_sembako_{tanggal}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name_excel}.xlsx"'},
    )

    variant_groceries = VariantGroceries.objects.all().order_by('-groceries_created')
    name_sembako = variant_groceries.values_list('groceries_name', flat=True)
    massa = variant_groceries.values_list('groceries_massa', flat=True)
    quantity = variant_groceries.values_list('groceries_quantity', flat=True)
    harga_hari_ini = sql_harga_hari_ini()
    harga_kemarin = sql_harga_kemarin()

    data = {
        'Nama Sembako': list(name_sembako),
        'Kuantitas': list(quantity),
        'Satuan': massa,
        'Harga Hari Ini': harga_hari_ini['price'],
        'Harga Kemarin': harga_kemarin
    }

    df = pd.DataFrame(data)

    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('Data')

        bold_format = workbook.add_format({'bold': True, 'font_size': 16})
        worksheet.write('A1', 'Harga Sembako', bold_format)

        worksheet.write('A2', f'Tanggal Export: {tanggal}')

        df.to_excel(writer, sheet_name='Data', startrow=3, index=False)

    return response


def sql_harga_hari_ini() -> dict:
    raw_query = """
    SELECT
        groceries_name, COALESCE(unit_groceries_price, 0) as price
    FROM
        administrator_variantgroceries av
    LEFT JOIN (
        SELECT
            unit_groceries_variant_id, MAX(unit_groceries_price) as unit_groceries_price
        FROM
            administrator_unitgroceries au
        WHERE DATE(au.unit_groceries_created) = CURDATE()
        GROUP BY unit_groceries_variant_id
        ) administrator_unitgroceries ON unit_groceries_variant_id = av.groceries_id;
    """

    with connection.cursor() as cursor:
        cursor.execute(raw_query)
        results = cursor.fetchall()

    data = {
        'price': [],
    }

    for row in results:
        data['price'].append(row[1])

    return data


def sql_harga_kemarin() -> list:
    raw_query = """
    SELECT
        groceries_name, COALESCE(unit_groceries_price, 0) as price
    FROM
        administrator_variantgroceries av
    LEFT JOIN (
        SELECT
            unit_groceries_variant_id, MAX(unit_groceries_price) as unit_groceries_price
        FROM
            administrator_unitgroceries au
        WHERE DATE(au.unit_groceries_created) = CURDATE() - INTERVAL 1 DAY
        GROUP BY unit_groceries_variant_id
        ) administrator_unitgroceries ON unit_groceries_variant_id = av.groceries_id;
    """

    with connection.cursor() as cursor:
        cursor.execute(raw_query)
        results = cursor.fetchall()

    price = []
    for row in results:
        price.append(row[1])

    return price


def import_excel_sembako(request):
    file = request.FILES.get('excel')
    tanggal = request.POST.get('tanggal')

    if not tanggal or not file:
        return JsonResponse({
            'status': False,
            'message': 'excel dan tanggal wajib diisi.'
        }, status=400)

    if file.size > 2_097_152:
        return JsonResponse({
            'status': False,
            'message': 'File tidak boleh lebih besar dari 2MB'
        }, status=400)

    if file.content_type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 'application/vnd.ms-excel']:
        return JsonResponse({
            'status': False,
            'message': 'File harus .xlsx dan .xls'
        }, status=400)

    file.read()
    df = pd.read_excel(file)
    if not check_format_excel(df):
        return JsonResponse({
            'status': False,
            'message': 'Format/Isi excel tidak valid.'
        }, status=400)

    for index, row in df.iterrows():
        obj, created = VariantGroceries.objects.get_or_create(groceries_name=row[0],
                                                              defaults={
                                                                  'groceries_massa': row[1],
                                                                  'groceries_quantity': row[2],
                                                              })
        UnitGroceries.objects.update_or_create(
            unit_groceries_variant_id=obj.pk,
            unit_groceries_created=tanggal,
            defaults={
                "unit_groceries_price": row[3],
            }
        )

    return JsonResponse({
        'success': True,
        'message': 'Berhasil import bahan pokok'
    })


def check_format_excel(data_frame: pd.DataFrame):
    valid_tipe: list = [str, str, int, int]

    list_valid = []
    for index, row in data_frame.iterrows():
        tipe = [type(i) for i in row]
        list_valid.append(tipe == valid_tipe)

    for valid in list_valid:
        if not valid:
            return valid

    return True


def export_ukm_to_excel(request):
    tanggal = datetime.now().strftime("%d/%B/%Y")
    name_excel = f"UKM_{tanggal}"

    # Query data dari model UKM
    ukm_data = UKM.objects.all()

    # Buat dataframe dari data model
    data = {
        'No': [ukm + 1 for ukm in range(0, ukm_data.count())],
        'Name': [ukm.ukm_name for ukm in ukm_data],
        'Owner': [ukm.ukm_owner for ukm in ukm_data],
        'Number Phone': [ukm.ukm_number_phone for ukm in ukm_data],
        'Legality': [ukm.ukm_legality for ukm in ukm_data],
        'Type of Product': [ukm.ukm_types_product for ukm in ukm_data],
        'Outlet': [ukm.ukm_outlet for ukm in ukm_data],
        'Address': [ukm.ukm_address for ukm in ukm_data],
        'Description': [ukm.ukm_description for ukm in ukm_data],
        'Website': [ukm.ukm_website for ukm in ukm_data],
        'Social Media': [ukm.ukm_social_media for ukm in ukm_data],
    }

    # Buat DataFrame dari data
    df = pd.DataFrame(data)

    # Buat response Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'fattachment; filename="{name_excel}.xlsx"'

    # Tulis data DataFrame ke response
    df.to_excel(response, index=False)

    return response


class ExportIKMToExcel(LoginRequiredMixin, View):
    def get(self, request):
        # Query semua data dari model IKM
        ikm_data = IKM.objects.all()

        # Buat DataFrame dari data IKM
        data = {
            'No': [ikm + 1 for ikm in range(0, ikm_data.count())],
            'Name': [ikm.ikm_name for ikm in ikm_data],
            'Owner': [ikm.ikm_owner for ikm in ikm_data],
            'Phone Number': [ikm.ikm_number_phone for ikm in ikm_data],
            'Legality': [ikm.ikm_legality for ikm in ikm_data],
            'Type of Product': [ikm.ikm_types_product for ikm in ikm_data],
            'Outlet': [ikm.ikm_outlet for ikm in ikm_data],
            'Address': [ikm.ikm_address for ikm in ikm_data],
            'Description': [ikm.ikm_description for ikm in ikm_data],
            'Website': [ikm.ikm_website for ikm in ikm_data],
            'Social Media': [ikm.ikm_social_media for ikm in ikm_data],
        }

        df = pd.DataFrame(data)

        tanggal = datetime.now().strftime("%d/%B/%Y")
        name_excel = f"Data_IKM_{tanggal}"

        # Buat response HTTP dengan file Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{name_excel}.xlsx"'

        # Tulis data DataFrame ke file Excel
        df.to_excel(response, index=False, engine='openpyxl')

        return response


class PasarView(ListView):
    model = Market
    template_name = 'visitor/pasar.html'
    extra_context = {
        'title': 'Pasar'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name_market')
        if name_filter:
            queryset = queryset.filter(market_name__icontains=name_filter)
        return queryset


class ExportPasarToExcel(LoginRequiredMixin, View):
    def get(self, request):
        # Query semua data dari model IKM
        market_data = Market.objects.all()

        # Buat DataFrame dari data IKM
        data = {
            'No': [index + 1 for index in range(0, market_data.count())],
            'Nama Pasar': [market.market_name for market in market_data],
            'Alamat Pasar': [market.market_address for market in market_data]
        }

        df = pd.DataFrame(data)

        tanggal = datetime.now().strftime("%d/%B/%Y")
        name_excel = f"Data_Pasar_{tanggal}"

        # Buat response HTTP dengan file Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{name_excel}.xlsx"'

        # Tulis data DataFrame ke file Excel
        df.to_excel(response, index=False, engine='openpyxl')

        return response


class AgenLPGView(ListView):
    template_name = 'visitor/agen_lpg.html'
    model = AgenLPG
    paginate_by = 12
    extra_context = {
        'title': 'Agen LPG'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(agen_name__icontains=name_filter)
        return queryset


class AgenLPGViewDetail(DetailView):
    model = AgenLPG
    template_name = 'visitor/agen-detail.html'
    slug_field = 'agen_slug'
    extra_context: dict = {}

    def get(self, request, *args, **kwargs):
        self.extra_context.update({'title': kwargs.get('slug')})
        return super().get(request)


class ExportAgenLPGToExcel(LoginRequiredMixin, View):
    def get(self, request):
        # Query semua data dari model IKM
        agen_data = AgenLPG.objects.all()

        # Buat DataFrame dari data IKM
        data = {
            'No': [index + 1 for index in range(0, agen_data.count())],
            'Nama Agen': [agen.agen_name for agen in agen_data],
            'Alamat Agen': [agen.agen_address for agen in agen_data],
            'Nama Pangkalan': [agen.agen_base_name for agen in agen_data],
            'Alamat Pangkalan': [agen.agen_base_address for agen in agen_data],
            'Nomor Telepon': [agen.agen_number_phone for agen in agen_data],
        }

        df = pd.DataFrame(data)

        tanggal = datetime.now().strftime("%d/%B/%Y")
        name_excel = f"Data_AgenLPG_{tanggal}"

        # Buat response HTTP dengan file Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{name_excel}.xlsx"'

        # Tulis data DataFrame ke file Excel
        df.to_excel(response, index=False, engine='openpyxl')

        return response


class KiosPupukListView(ListView):
    model = KiosPupuk
    template_name = 'visitor/kios_pupuk.html'
    paginate_by = 12
    extra_context = {
        'title': 'Kios Pupuk'
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(kios_name__icontains=name_filter)
        return queryset


class KiosPupukDetailView(DetailView):
    model = KiosPupuk
    template_name = 'visitor/kios_pupuk_detail.html'
    slug_field = 'kios_slug'
    extra_context: dict = {}

    def get(self, request, *args, **kwargs):
        self.extra_context.update({'title': kwargs.get('slug')})
        return super().get(request)


class ExportKiosPupukToExcel(LoginRequiredMixin, View):
    def get(self, request):
        # Query semua data dari model IKM
        kios_data = KiosPupuk.objects.all()

        # Buat DataFrame dari data IKM
        data = {
            'No': [index + 1 for index in range(0, kios_data.count())],
            'Nama Kios': [kios.kios_name for kios in kios_data],
            'Alamat Kios': [kios.kios_address for kios in kios_data],
            'Nomor Telepon': [kios.kios_number_phone for kios in kios_data],
            'Nama Distributor': [kios.kios_distributor for kios in kios_data],
            'Alamat Distributor': [kios.kios_distributor_address for kios in kios_data],
        }

        df = pd.DataFrame(data)

        tanggal = datetime.now().strftime("%d/%B/%Y")
        name_excel = f"Data_KiosData_{tanggal}"

        # Buat response HTTP dengan file Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{name_excel}.xlsx"'

        # Tulis data DataFrame ke file Excel
        df.to_excel(response, index=False, engine='openpyxl')

        return response


class StockItemListView(ListView):
    model = StockItem
    template_name = 'visitor/stock_item.html'
    extra_context = {
        'title': 'Stok Barang'
    }
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        if name_filter:
            queryset = queryset.filter(item_name__icontains=name_filter)
        return queryset


class ExportStockItemToExcel(LoginRequiredMixin, View):
    def get(self, request):
        # Query semua data dari model IKM
        stock_data = StockItem.objects.all()

        # Buat DataFrame dari data IKM
        data = {
            'No': [index + 1 for index in range(0, stock_data.count())],
            'Nama Barang': [stock.item_name for stock in stock_data],
            'Nama Pasar': [stock.item_market for stock in stock_data],
            'Stok Awal': [stock.item_stock for stock in stock_data],
            'Barang Masuk': [stock.item_income for stock in stock_data],
            'Penjualan': [stock.item_outcome for stock in stock_data],
            'Stok Akhir': [stock.item_last_stock for stock in stock_data],
        }

        df = pd.DataFrame(data)

        tanggal = datetime.now().strftime("%d/%B/%Y")
        name_excel = f"Data_StokBarang_{tanggal}"

        # Buat response HTTP dengan file Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{name_excel}.xlsx"'

        # Tulis data DataFrame ke file Excel
        df.to_excel(response, index=False, engine='openpyxl')

        return response


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)


def dataForStatisticSembako(request):
    today = datetime.today().date()
    one_week_ago = today - timedelta(days=7)

    id_sembako = request.GET.get('id_sembako')
    start_date_str = request.GET.get('start_date', one_week_ago.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))

    if not id_sembako:
        return JsonResponse({'success': False, 'error': 'id_sembako harus diisi'})

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    result_data = []

    for single_date in daterange(start_date, end_date):
        try:
            query = UnitGroceries.objects.get(
                unit_groceries_variant_id=id_sembako,
                unit_groceries_created=single_date,
            )
            result_data.append({
                'unit_groceries_price': query.unit_groceries_price,
                'unit_groceries_created': query.unit_groceries_created
            })
        except UnitGroceries.DoesNotExist:
            result_data.append({
                'unit_groceries_price': 0,
                'unit_groceries_created': single_date
            })

    return JsonResponse({
        'success': True,
        'data': result_data
    })


def allVariantSembako(request):
    query = VariantGroceries.objects.all().values('groceries_name', 'groceries_id')

    return JsonResponse({
        'success': True,
        'data': list(query)
    })


def export_groceries_data(request):
    today = date.today().strftime('%d/%B/%Y')
    title = INVALID_TITLE_REGEX.sub('_', f"Harga_sembako_{today}")

    # Membuat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Menambahkan judul "Tambah Sembako" dan tanggal ekspor
    ws.append(["Harga Sembako"])
    ws.append([f"Tanggal Export: {today}"])
    ws.append([])  # Untuk memberi jarak antara judul dan data

    # Menuliskan header
    header = ["Nama Sembako", "Kuantitas","Satuan", "Harga Hari Ini", "Harga Kemarin"]
    ws.append(header)

    # Mengatur format header menjadi tebal (bold)
    for cell in ws[1]:
        cell.font = Font(bold=True, sz=16)

    # Mendapatkan tanggal kemarin
    yesterday = date.today() - timedelta(days=1)

    # Mendapatkan semua data VariantGroceries
    variant_groceries = VariantGroceries.objects.all()

    # Mengisi data ke dalam file Excel
    for variant in variant_groceries:
        # Mengambil harga kemarin
        yesterday_price = variant.unit_groceries.filter(unit_groceries_created=yesterday).first()
        yesterday_price = yesterday_price.unit_groceries_price if yesterday_price else 0

        # Mengambil harga hari ini
        today_price = variant.unit_groceries.filter(unit_groceries_created=date.today()).first()
        today_price = today_price.unit_groceries_price if today_price else 0

        ws.append([variant.groceries_name, variant.groceries_quantity,variant.groceries_massa, today_price, yesterday_price])

    # Mengatur data menjadi tabel
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(top=Side(style='thin'),
                                 right=Side(style='thin'),
                                 bottom=Side(style='thin'),
                                 left=Side(style='thin'))

    # Menggunakan HttpResponse untuk mengirim file Excel ke client
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={title}.xlsx'
    wb.save(response)
    return response
